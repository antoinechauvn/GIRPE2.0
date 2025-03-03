import os
import shutil
import win32com.client
from PySide6.QtCore import QObject, QThread
from openpyxl import load_workbook
from jinja2 import Template, Environment, StrictUndefined


class ConversionThread(QThread):
    """Thread pour la conversion Excel vers PDF."""

    def __init__(self, excel_path, pdf_path):
        super().__init__()
        self.excel_path = excel_path
        self.pdf_path = pdf_path

    def run(self):
        """Convertit le fichier Excel en PDF."""
        excel = None
        wb = None
        try:
            print("Conversion Excel vers PDF en cours...")
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            
            # Utiliser des chemins absolus
            abs_excel_path = os.path.abspath(self.excel_path)
            abs_pdf_path = os.path.abspath(self.pdf_path)
            
            wb = excel.Workbooks.Open(abs_excel_path)
            wb.ExportAsFixedFormat(
                Type=0,  # PDF
                Filename=abs_pdf_path,
                Quality=0,  # Standard
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            print("Conversion PDF terminée")
        except Exception as e:
            print(f"Erreur pendant la conversion: {str(e)}")
            raise
        finally:
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.EnableEvents = True
                    excel.Quit()
                    del excel
            except:
                pass  # Ignorer les erreurs lors du nettoyage


class ExcelToPdfWorker(QObject):
    """Worker pour convertir un fichier Excel en PDF."""

    def __init__(self):
        super().__init__()
        self.excel_path = None
        self.pdf_path = None
        self.template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "resource", "feuille-equipes-4-joueurs.xlsx")
        self.thread = None

    def set_paths(self, filename: str):
        """
        Définit les chemins des fichiers Excel et PDF.
        
        Args:
            filename: Nom du fichier sans extension
        """
        self.excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "matches", f"{filename}.xlsx")
        self.pdf_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "matches", f"{filename}.pdf")

    def create_excel_file(self):
        """Crée une copie du template Excel."""
        # Créer le dossier matches s'il n'existe pas
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
        
        # Copier le template
        shutil.copy2(self.template_path, self.excel_path)

    def create_thread(self):
        """Crée et retourne un nouveau thread de conversion."""
        return ConversionThread(self.excel_path, self.pdf_path)

    def convert(self):
        """Lance la conversion dans un thread séparé.
        
        Returns:
            ConversionThread: Le thread créé mais non démarré, ou None si une conversion est déjà en cours
        """
        if self.thread and self.thread.isRunning():
            print("Une conversion est déjà en cours")
            return None
            
        self.thread = self.create_thread()
        return self.thread  # Retourne le thread sans le démarrer

    @staticmethod
    def update_variables(excel_path: str, variables: dict):
        """
        Met à jour les variables jinja dans le fichier Excel.
        
        Args:
            excel_path: Chemin vers le fichier Excel à mettre à jour
            variables: Dictionnaire des variables à remplacer
        """
        wb = load_workbook(excel_path)
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    value = cell.value
                    for var_name, var_value in variables.items():
                        # Remplacer la variable si elle existe dans la cellule
                        value = value.replace('{{ ' + var_name + ' }}', str(var_value))
                        value = value.replace('{{' + var_name + '}}', str(var_value))
                    cell.value = value

        wb.save(excel_path)
        wb.close()  # Fermer le workbook après avoir sauvegardé
